#!/usr/bin/env python3
r"""
Creo_CSV2XLSX.py

Konvertiert eine hierarchische Creo-Stückliste (CSV) in eine strukturierte
Excel-Arbeitsmappe.

Änderungshistorie (neueste zuerst):
  V1.0.3  05.08.26   XLSX-Dateiname richtet sich jetzt nach der Sachnummer aus
                     Ebene E1 (z.B. '38-02024.00_STUECKLISTE.XLSX'), unabhängig
                     vom Namen der CSV-Quelldatei. Schlosserliste-Blatt und PDF
                     heißen entsprechend '..._SCHLOSSERLISTE' (Blattname und
                     PDF-Dateiname sind identisch). Die Erkennung, ob eine
                     bestehende XLSX fortgesetzt wird (Modus FORTSETZEN), läuft
                     jetzt ebenfalls über diese Archivnummer statt über den
                     CSV-Dateinamen. Neue Häkchen-Spalte (Spalte A) in der
                     Schlosserliste zum manuellen Abhaken erledigter Positionen.

  V1.0.2  04.08.26   Wärmebehandlung-Mapping erweitert (Plasmanitriert,
                     Tiefennitriert, Brüniert, Vakuumhärten) und robuster
                     gemacht: Schreibweisen mit Umlaut oder ae/oe/ue/ss sowie
                     Groß-/Kleinschreibung werden jetzt einheitlich über eine
                     Normalisierungsfunktion erkannt (vorher nur .lower()).
                     Einzelgewicht wird bei zusammengeführten Positionen nicht
                     mehr aufsummiert (bleibt Stückgewicht der Einzelposition);
                     Masse-Formel im Tabellenkopf der Schlosserliste korrigiert
                     auf SUMPRODUCT(Stückzahl, Einzelgewicht) statt reiner
                     Summe der Gewichtsspalte.

  V1.0.1  03.08.26   Kumulierung der Unterbaugruppen-Stückzahl korrigiert -
                     eine Baugruppe, die selbst mehrfach verbaut ist, hat ihre
                     Stückzahl bisher nicht an ihre Unterteile weitergegeben
                     (Multiplikator-Kette für verschachtelte Baugruppen ergänzt).
                     Prüfhinweise-Spalte im Stand-Blatt hinzugefügt (Logging
                     direkt in der Excel-Ausgabe, betroffene Zeilen gelb markiert).

  Ältere Änderungen bitte hier oberhalb der jeweils vorherigen ergänzen, mit
  Versionsnummer, Datum und einer kurzen Beschreibung - siehe Format oben.

Wichtig (Blatt-Reihenfolge):
  1) Tabelle1  → <Auftragsnummer>_fuer_GFU
  2) Tabelle2  → <Auftragsnummer>_Schlosserliste
  3) Tabelle3  → aktuellstes Stücklisten-Stand-Blatt (<stamm>_Stand_<YYYYMMDD>)
  4) danach    → alle älteren Stand-Blätter

Aufruf:
    python Creo_CSV2XLSX.py "C:\Pfad\zum\Ordner" [--batch]

Im Ordner darf genau eine CSV liegen (beliebiger Name).

Wenn bereits eine XLSX zum gleichen Auftragsstamm existiert:
  - GFU- und Schlosserliste werden gelöscht und neu erzeugt
  - ein neues Stand-Blatt wird angehängt (Versionsstände bleiben erhalten)

Am Ende wird die CSV gelöscht (alle Daten sind in der XLSX gesichert).

Schweißteil-Erkennung (strukturell):
  Eine Baugruppe ist ein Schweißteil, wenn ALLE direkten Kinder Blattteile
  vom Typ "Brennteil" sind. Bereits korrekt geführte Schweißteile werden
  automatisch bestätigt. Neue Kandidaten werden interaktiv abgefragt.

Abhängigkeiten:
  openpyxl (pip install openpyxl)
  optionaler Zipimport-Fallback: N:\Download\PortableApps\Python\openpyxl.zip

Exit code:
  0 = OK, 1 = Fehler
"""

import time
from datetime import date
from pathlib import Path

# Versionsnummer manuell hier UND in der Änderungshistorie oben pflegen -
# ein automatischer Zeitstempel (z.B. aus dem Datei-Änderungsdatum) sagt
# nichts darüber aus, WAS sich geändert hat.
SCRIPT_VERSION = "V1.0.2"

import argparse
import csv
import re
import subprocess
import sys

# Lazy import openpyxl
openpyxl = None
Alignment = Font = PatternFill = Border = Side = None
get_column_letter = None
PageMargins = None
HDR_FILL_BLAU = HDR_FILL_GRUEN = SW_FILL = HDR_FONT = BORDER = None
GRAU_FILL = None
HINWEIS_FILL = None

# Schlosserliste-Vorlagenstyles (werden nach openpyxl-Import gesetzt)
HDR_FILL_ROT = None
SCHLOSSER_FONT = None
SCHLOSSER_HDR_FONT = None


def _import_openpyxl() -> None:
    """Importiert openpyxl verzögert. Fallback per Zipimport falls nicht installiert."""
    global openpyxl, Alignment, Font, PatternFill, Border, Side, get_column_letter, PageMargins
    global HDR_FILL_BLAU, HDR_FILL_GRUEN, SW_FILL, HDR_FILL_ROT, HDR_FONT, BORDER
    global SCHLOSSER_FONT, SCHLOSSER_HDR_FONT
    global GRAU_FILL, HINWEIS_FILL

    if openpyxl is not None:
        return

    t0 = time.time()
    try:
        import openpyxl as _openpyxl
    except ImportError:
        zip_pfad = Path(r'N:\Download\PortableApps\Python\openpyxl.zip')
        if not zip_pfad.exists():
            print(f"[X]   openpyxl ist nicht installiert und '{zip_pfad}' wurde nicht gefunden.")
            raise
        print(f"[>]   openpyxl nicht installiert - lade aus '{zip_pfad}' (Zipimport)")
        sys.path.insert(0, str(zip_pfad))
        import openpyxl as _openpyxl

    from openpyxl.styles import (
        Alignment as _Alignment, Font as _Font, PatternFill as _PatternFill,
        Border as _Border, Side as _Side,
    )
    from openpyxl.utils import get_column_letter as _get_column_letter
    from openpyxl.worksheet.page import PageMargins as _PageMargins

    openpyxl = _openpyxl
    Alignment, Font, PatternFill, Border, Side = _Alignment, _Font, _PatternFill, _Border, _Side
    get_column_letter = _get_column_letter
    PageMargins = _PageMargins

    HDR_FILL_BLAU = PatternFill('solid', fgColor='BDD7EE')
    HDR_FILL_GRUEN = PatternFill('solid', fgColor='C6EFCE')
    HDR_FILL_ROT = PatternFill('solid', fgColor='C00000')
    SW_FILL = PatternFill('solid', fgColor='FCE4D6')
    GRAU_FILL = PatternFill('solid', fgColor='D9D9D9')
    HINWEIS_FILL = PatternFill('solid', fgColor='FFF2CC')

    HDR_FONT = Font(bold=True)
    SCHLOSSER_FONT = Font(name='Bahnschrift SemiCondensed', sz=16, bold=False)
    SCHLOSSER_HDR_FONT = Font(name='Bahnschrift SemiCondensed', sz=16, bold=True)

    _thin = Side(style='thin')
    BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── Konstanten ────────────────────────────────────────────────────────────────

EBENE_PATTERN = re.compile(r'^E(\d+)$')
HEADER_SEARCH_ROWS = 15

COLUMN_NAMES = [
    "Lfd.", "Stück", "Benennung", "Sachnummer", "Pos-Nr", "Ebene",
    "Norm-Kurzbezeichnung", "Material", "Gewicht", "Wärmebehandlung",
    "Bemerkungen", "F,K,N", "V", "Rev.", "Dateiname", "Lieferant",
]
BASIS_SPALTEN = ["Stück", "Benennung", "Sachnummer", "Pos-Nr", "Ebene"]

FLAT_COLUMNS = [
    "Lfd.", "Stück", "Benennung", "Sachnummer", "Pos-Nr",
    "Norm-Kurzbezeichnung", "Material", "Gewicht", "Wärmebehandlung",
    "Bemerkungen", "F,K,N", "V", "Rev.", "Dateiname", "Lieferant",
]

def normalisiere_waerme_schluessel(s: str) -> str:
    """Vereinheitlicht Schreibweisen für den Wärmebehandlung-Abgleich:
    Kleinschreibung + Umlaute auf ae/oe/ue/ss (ASCII) vereinheitlicht.
    So matchen 'glühen', 'Glühen' und 'gluehen' alle auf denselben Schlüssel,
    unabhängig davon, wie Creo den Text im jeweiligen Export schreibt."""
    s = s.strip().lower()
    for a, b in (('ä', 'ae'), ('ö', 'oe'), ('ü', 'ue'), ('ß', 'ss')):
        s = s.replace(a, b)
    return s


# Schlüssel bereits normalisiert (siehe normalisiere_waerme_schluessel) - beim
# Ergänzen neuer Einträge normale/korrekte Schreibweise nehmen, hier unten
# in Kleinbuchstaben mit ae/oe/ue/ss eintragen, der Rest passiert automatisch.
WAERME_MAP = {
    'einsatzgehaertet': 'TE+TH',
    'verguetet': 'VERGUE',
    'nitriert': 'NIT',
    'plasmanitriert': 'NIT-P',
    'tiefennitriert': 'NIT-T',
    'brueniert': 'BRUE',
    'gluehen': 'GLUE',
    'vakuumhaerten': 'VAK-H',
    'vakuumgehaertet': 'VAK-H',  # ältere/alternative Schreibweise, gleiches Kürzel
}

SCHWEISSTEIL_RE = re.compile(r'^schwei[sß]s?teil$', re.IGNORECASE)

# Schlosserliste: wie Vorlage 38-02024-00_Schlosserliste
SCHLOSSER_HEADERS = [
    "✓",
    "Pos.", "Stk.",
    "Benennung\n(Dateiname)",
    "Sachnummer\nNorm-Kurzbezeichnung",
    "Material\nWärmebeh.",
    "Masse\n0kg",  # wird als Formel überschrieben
    "Lieferant\nBemerkungen",
    "F,K,N", "V", "Rev.",
]
SCHLOSSER_BREITEN = [6.0, 7.7109375, 5.28515625, 28.7109375, 32.7109375, 21.7109375,
                     10.7109375, 17.7109375, 4.7109375, 5.0, 6.0]
SCHLOSSER_MERGES = ["D2:D3", "E2:E3", "F2:G3", "H2:H3", "I2:K3", "F4:G5", "H4:H5"]


# ── Logging ───────────────────────────────────────────────────────────────────

_LEVEL_SYMBOL = {'INFO': '>', 'WARN': '!', 'FEHLER': 'X'}
ZEILEN_BREITE = 3  # wird in process() an die tatsächliche Zeilenzahl angepasst


def log(level: str, msg: str, zeile=None, benennung=None, sachnummer=None, posnr=None) -> None:
    """Gibt eine Meldung aus: [Symbol]Zeile NNN: Benennung (Sachnummer) Pos X: msg
    Symbol: '>' = INFO, '!' = WARN, 'X' = FEHLER. Ohne Zeile: [Symbol] msg."""
    symbol = _LEVEL_SYMBOL.get(level, level)
    if zeile is None:
        print(f"[{symbol}] {msg}")
        return
    teil = str(benennung or '').strip()
    if sachnummer:
        teil += f" ({sachnummer})"
    if posnr:
        teil = (teil + " " if teil else "") + f"Pos {posnr}"
    praefix = f"Zeile {zeile:>{ZEILEN_BREITE}}: " + (f"{teil}: " if teil else "")
    print(f"[{symbol}]{praefix}{msg}")


# ── Hilfsfunktionen ───────────────────────────────────────────────────────────

def normalize_schweissteil(value):
    if value and SCHWEISSTEIL_RE.match(str(value).strip()):
        return 'Schweißteil'
    return value


def convert_waermebehandlung(raw, entry):
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ('', '-', ' -'):
        return s
    mapped = WAERME_MAP.get(normalisiere_waerme_schluessel(s))
    if mapped:
        return mapped
    log('WARN', f"Wärmebehandlung '{s}' nicht im Mapping – Originalwert behalten",
        zeile=entry['row'], benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
    return s


def parse_gewicht(raw, entry):
    if raw is None or str(raw).strip() in ('', '-', ' -'):
        return None
    s = str(raw).strip()
    if '.' in s:
        log('WARN', f"Gewicht '{s}' enthält '.' – möglicher Lokalisierungsfehler",
            zeile=entry['row'], benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
    try:
        val = float(s.replace(',', '.'))
    except ValueError:
        log('WARN', f"Gewicht '{s}' nicht lesbar – Zelle bleibt leer",
            zeile=entry['row'], benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
        return None
    if val == 0.0:
        log('WARN', "Gewicht = 0 ist unplausibel",
            zeile=entry['row'], benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
    return val


def auto_col_width(ws, min_width: int = 8, max_width: int = 55, padding: int = 2) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is None:
                continue
            lines = str(cell.value).split('\n')
            cell_len = max(len(l) for l in lines)
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max(max_len + padding, min_width), max_width)


def copy_titelblock(src_ws, dst_ws, header_row: int, max_col: int) -> None:
    for r in range(1, header_row):
        for c in range(1, max_col + 1):
            val = src_ws.cell(r, c).value
            if val is not None:
                dst_ws.cell(r, c, val)
        dst_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height or 15


# ── CSV / Workbook einlesen ───────────────────────────────────────────────────

def load_csv(path: Path):
    _import_openpyxl()
    with open(path, encoding='utf-8-sig', newline='') as fh:
        raw_rows = list(csv.reader(fh))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'tmp'
    for r_idx, row in enumerate(raw_rows, 1):
        for c_idx, val in enumerate(row, 1):
            if isinstance(val, str):
                val = val.strip() or None
            ws.cell(r_idx, c_idx, val)
    return wb, ws, raw_rows


def find_header(ws):
    for r in range(1, min(ws.max_row, HEADER_SEARCH_ROWS) + 1):
        hmap = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() in COLUMN_NAMES:
                hmap[v.strip()] = c
        if all(col in hmap for col in BASIS_SPALTEN):
            return r, hmap
    raise ValueError(f"Kopfzeile mit Spalten {BASIS_SPALTEN} nicht in den ersten {HEADER_SEARCH_ROWS} Zeilen gefunden")


def find_auftragsnummer(ws, header_row: int) -> str:
    """Findet die Sachnummer der Ebene E1 (Punktformat, z.B. '38-02024.00') im
    Titelblock. Das ist die kanonische Kennung für Datei- und Blattnamen -
    unabhängig davon, wie die CSV-Quelldatei selbst benannt ist."""
    pattern = re.compile(r'\d{2}-\d{5}\.\d{2}')
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or '')
            m = pattern.search(v)
            if m:
                return m.group(0)
    return 'unbekannt'


def read_rows(ws, header_row: int, hmap: dict) -> list:
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        e_val = ws.cell(r, hmap['Ebene']).value
        if not e_val or str(e_val).strip() == '':
            continue
        entry = {'row': r}
        for name, col in hmap.items():
            entry[name] = ws.cell(r, col).value
        m = EBENE_PATTERN.match(str(e_val).strip())
        if not m:
            entry.update({'level': None, 'raw_ebene': e_val, 'error': 'Format ungültig'})
        else:
            entry['level'] = int(m.group(1))
            entry['teilenummer'] = str(entry.get('Pos-Nr') or '').strip()
            entry['error'] = None
            entry['Gewicht'] = parse_gewicht(entry.get('Gewicht'), entry)
            entry['Wärmebehandlung'] = convert_waermebehandlung(entry.get('Wärmebehandlung'), entry)
            nk_orig = entry.get('Norm-Kurzbezeichnung')
            nk_new = normalize_schweissteil(nk_orig)
            if nk_new != nk_orig:
                log('WARN', f"Norm-Kurzbezeichnung '{nk_orig}' → korrigiert zu 'Schweißteil'",
                    zeile=r, benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
            entry['Norm-Kurzbezeichnung'] = nk_new
            try:
                entry['Stück'] = int(entry.get('Stück') or 0)
            except (ValueError, TypeError):
                pass
        rows.append(entry)
    return rows


# ── Baumaufbau ────────────────────────────────────────────────────────────────

def build_tree(rows: list):
    valid = [e for e in rows if not e.get('error')]
    children = {i: [] for i in range(len(valid))}
    parent_of = {}
    stack = []
    for i, entry in enumerate(valid):
        while stack and stack[-1][0] >= entry['level']:
            stack.pop()
        parent_of[i] = stack[-1][1] if stack else None
        if stack:
            children[stack[-1][1]].append(i)
        stack.append((entry['level'], i))
    return valid, children, parent_of


# ── Validierung ───────────────────────────────────────────────────────────────

def validate_all(rows: list, valid: list, children: dict) -> list:
    problems = []
    prev_level = None
    for entry in rows:
        b, s, p = entry.get('Benennung'), entry.get('Sachnummer'), entry.get('Pos-Nr')
        if entry.get('error'):
            problems.append((entry['row'], 'FEHLER', f"Ungültiges Ebenen-Format: {entry.get('raw_ebene')!r}", b, s, p))
            continue
        level = entry['level']
        if prev_level is not None and level > prev_level + 1:
            problems.append((entry['row'], 'FEHLER', f"Ebenensprung von {prev_level} auf {level}", b, s, p))
        prev_level = level

        if not str(entry.get('Pos-Nr') or '').strip():
            problems.append((entry['row'], 'FEHLER', "Pos-Nr ist leer", b, s, p))
        if not entry.get('Benennung'):
            problems.append((entry['row'], 'FEHLER', "Benennung ist leer", b, s, p))
        if entry.get('Stück') in (None, ''):
            problems.append((entry['row'], 'FEHLER', "Stück ist leer", b, s, p))
        if not str(entry.get('Sachnummer') or '').strip():
            problems.append((entry['row'], 'WARN', "Sachnummer fehlt", b, s, p))

        lieferant = str(entry.get('Lieferant') or '').strip()
        if lieferant and lieferant != lieferant.upper():
            problems.append((entry['row'], 'WARN', f"Lieferant '{lieferant}' ist nicht vollständig großgeschrieben", b, s, p))

    by_posnr = {}
    for i, entry in enumerate(valid):
        if children[i]:
            continue
        key = str(entry.get('Pos-Nr') or '').strip()
        by_posnr.setdefault(key, []).append(entry)
    for posnr, group in by_posnr.items():
        if len(group) < 2:
            continue
        ref = group[0]
        for other in group[1:]:
            unterschiedlich = any(
                str(ref.get(f)) != str(other.get(f)) for f in ('Benennung', 'Sachnummer', 'Gewicht')
            )
            if unterschiedlich:
                problems.append((
                    other['row'], 'WARN', f"gleiche Pos-Nr vorhanden (bereits Zeile {ref['row']})",
                    other.get('Benennung'), other.get('Sachnummer'), other.get('Pos-Nr'),
                ))

    b_pattern = re.compile(r'^B\d{2}$')
    bg_pattern = re.compile(r'(?i)^BG\d+$')
    for i, entry in enumerate(valid):
        if not children[i]:
            continue
        posnr = str(entry.get('Pos-Nr') or '').strip()
        if posnr.startswith('E'):
            continue
        b, s = entry.get('Benennung'), entry.get('Sachnummer')
        if bg_pattern.match(posnr):
            problems.append((entry['row'], 'WARN', f"verwendet 'BG'-Präfix – Konvention ist 'Bxx' (z.B. 'B{posnr[2:].zfill(2)}')", b, s, posnr))
        elif not b_pattern.match(posnr):
            problems.append((entry['row'], 'WARN', "entspricht nicht dem Muster 'Bxx' (B + 2 Ziffern)", b, s, posnr))

    problems.sort(key=lambda p: p[0])
    return problems


# ── Schweißteil-Erkennung ─────────────────────────────────────────────────────

def bemerkung_besteht_aus(valid: list, kinder: list) -> str:
    return 'besteht aus Teil ' + ', '.join(str(valid[c].get('teilenummer')) for c in kinder)


def find_und_resolve_schweissteil(valid: list, children: dict, interaktiv: bool = True) -> set:
    schweissteil_rows: set = set()
    for i, entry in enumerate(valid):
        ch = children[i]
        if not ch:
            continue
        all_brennteil = all((not children[c]) and str(valid[c].get('Norm-Kurzbezeichnung') or '').strip().lower() == 'brennteil' for c in ch)
        if not all_brennteil:
            continue

        bemerkung = bemerkung_besteht_aus(valid, ch)
        cur = normalize_schweissteil(entry.get('Norm-Kurzbezeichnung'))

        if cur == 'Schweißteil':
            entry['Norm-Kurzbezeichnung'] = 'Schweißteil'
            entry['Bemerkungen'] = bemerkung
            schweissteil_rows.add(entry['row'])
            log('INFO', f"Schweißteil bestätigt – {bemerkung}", zeile=entry['row'],
                benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
        elif interaktiv:
            kinder_namen = [str(valid[c].get('Benennung')) for c in ch]
            print(f"\n  Baugruppe {entry.get('Pos-Nr')} '{entry.get('Benennung')}' (Zeile {entry['row']}) "
                  f"wurde als mögliches Schweißteil erkannt,")
            print(f"  es wurde aber keine Norm-Kurzbezeichnung gefunden "
                  f"(besteht nur aus Brennteilen: {', '.join(kinder_namen)}).")
            antwort = input("  Als 'Schweißteil' führen? [j/n]: ").strip().lower()
            print()
            if antwort == 'j':
                entry['Norm-Kurzbezeichnung'] = 'Schweißteil'
                entry['Bemerkungen'] = bemerkung
                schweissteil_rows.add(entry['row'])
                log('INFO', "→ als Schweißteil geführt", zeile=entry['row'],
                    benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
            else:
                log('INFO', "→ unverändert (Kinder einzeln in GFU)", zeile=entry['row'],
                    benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))
        else:
            entry['Norm-Kurzbezeichnung'] = 'Schweißteil'
            entry['Bemerkungen'] = bemerkung
            schweissteil_rows.add(entry['row'])
            log('INFO', f"Schweißteil (auto) – {bemerkung}", zeile=entry['row'],
                benennung=entry.get('Benennung'), sachnummer=entry.get('Sachnummer'), posnr=entry.get('Pos-Nr'))

    return schweissteil_rows


# ── GFU-Positionen sammeln & kumulieren ──────────────────────────────────────

def collect_items(valid: list, children: dict, parent_of: dict, schweissteil_rows: set):
    """Sammelt die tatsächlichen GFU-/Schlosser-Positionen. Wichtig: die Stückzahl
    einer Baugruppe multipliziert sich mit der Stückzahl ALLER Elternbaugruppen
    (BOM-Explosion) - eine Schraube, die 4x in einer Baugruppe verbaut ist, die
    selbst 2x in der Vorrichtung vorkommt, ergibt in Summe 8 Stück, nicht 4."""
    schweissteile: list = []
    sonstige: list = []

    def eigene_stueckzahl(entry: dict) -> int:
        try:
            return int(entry.get('Stück') or 1)
        except (TypeError, ValueError):
            return 1

    def visit(i: int, multiplikator: int) -> None:
        entry = valid[i]
        ch = children[i]
        gesamt_multiplikator = multiplikator * eigene_stueckzahl(entry)
        if not ch or entry['row'] in schweissteil_rows:
            kopie = dict(entry)
            kopie['Stück'] = gesamt_multiplikator
            (schweissteile if entry['row'] in schweissteil_rows else sonstige).append(kopie)
        else:
            for c in ch:
                visit(c, gesamt_multiplikator)

    roots = [i for i in range(len(valid)) if parent_of[i] is None]
    for r in roots:
        visit(r, 1)
    return schweissteile, sonstige


def aggregate_items(schweissteile: list, sonstige: list):
    """Kumuliert nach (Sachnummer, Pos-Nr) zusammen - das ist immer eindeutig:
    - gleiche Sachnummer + gleiche Pos-Nr = wirklich dasselbe Teil (z.B. dieselbe
      Norm-Schraube in mehreren Unterbaugruppen) -> Stück wird summiert.
    - gleiche Pos-Nr, aber unterschiedliche Sachnummer (z.B. Creo hat eine
      Pos-Nr fälschlich doppelt vergeben) -> bleiben getrennte Zeilen, da der
      Schlüssel unterschiedlich ist. Die Validierung meldet den Pos-Nr-Konflikt
      trotzdem als Hinweis (siehe validate_all)."""
    def kumulate(items: list) -> list:
        groups: dict = {}
        for entry in items:
            sachnr = str(entry.get('Sachnummer') or '').strip()
            posnr = str(entry.get('Pos-Nr') or '').strip()
            key = (sachnr, posnr)

            if key not in groups:
                groups[key] = dict(entry)
            else:
                try:
                    groups[key]['Stück'] = int(groups[key].get('Stück') or 0) + int(entry.get('Stück') or 0)
                except (TypeError, ValueError) as e:
                    log('WARN', f"Stück-Summierung fehlgeschlagen ({e})", zeile=entry.get('row'),
                        benennung=entry.get('Benennung'), sachnummer=sachnr, posnr=posnr)
                # Gewicht bleibt das Einzelgewicht (Stückgewicht) der ersten
                # Fundstelle - wird NICHT aufsummiert. Die Gesamtmasse wird nur
                # im Tabellenkopf der Schlosserliste kumuliert (Summenformel).
                log('INFO', f"kumuliert – {groups[key]['Stück']} Stück gesamt", zeile=entry.get('row'),
                    benennung=entry.get('Benennung'), sachnummer=sachnr, posnr=posnr)

        def sort_key(entry):
            try:
                return (0, int(str(entry.get('Pos-Nr') or '')))
            except ValueError:
                return (1, str(entry.get('Pos-Nr') or ''))

        return sorted(groups.values(), key=sort_key)

    result_sw = kumulate(schweissteile)
    result_so = kumulate(sonstige)
    alle = result_sw + result_so
    for idx, r in enumerate(alle, 1):
        r['Lfd.'] = idx
    return alle, len(result_sw)


# ── Blatt schreiben: Quellblatt ───────────────────────────────────────────────

def write_source_sheet(wb, src_ws, header_row: int, hmap: dict, valid: list, sheet_name: str,
                       probleme_je_zeile: dict = None):
    ws = wb.create_sheet(sheet_name)
    n_col = max(hmap.values())
    hinweis_col = n_col + 1
    copy_titelblock(src_ws, ws, header_row, n_col)

    for name, col in hmap.items():
        cell = ws.cell(header_row, col, name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL_BLAU
        cell.border = BORDER
    hcell = ws.cell(header_row, hinweis_col, "Prüfhinweise")
    hcell.font = HDR_FONT
    hcell.fill = HDR_FILL_BLAU
    hcell.border = BORDER
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(hinweis_col)}{header_row}"

    probleme_je_zeile = probleme_je_zeile or {}
    by_row = {e['row']: e for e in valid}
    for src_row, entry in by_row.items():
        for name, col in hmap.items():
            cell = ws.cell(src_row, col, entry.get(name))
            cell.border = BORDER
        level = entry.get('level', 1)
        b_col = hmap.get('Benennung')
        if b_col:
            ws.cell(src_row, b_col).alignment = Alignment(indent=(level - 1) * 2)
        ws.row_dimensions[src_row].outline_level = min(level - 1, 7)

        meldungen = probleme_je_zeile.get(src_row)
        hz = ws.cell(src_row, hinweis_col)
        hz.border = BORDER
        if meldungen:
            hz.value = '; '.join(meldungen)
            hz.alignment = Alignment(wrap_text=True, vertical='top')
            for c in range(1, hinweis_col + 1):
                ws.cell(src_row, c).fill = HINWEIS_FILL
        else:
            hz.value = None

    ws.sheet_properties.outlinePr.summaryBelow = False
    auto_col_width(ws)
    return ws


# ── Blatt schreiben: GFU-Flachliste ──────────────────────────────────────────

def write_flat_sheet(wb, src_ws, header_row: int, aggregated: list, n_schweissteile: int, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    n_col = len(FLAT_COLUMNS)
    copy_titelblock(src_ws, ws, header_row, n_col)

    for c, name in enumerate(FLAT_COLUMNS, 1):
        cell = ws.cell(header_row, c, name)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL_BLAU
        cell.border = BORDER
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_col)}{header_row}"

    for r_off, entry in enumerate(aggregated, header_row + 1):
        is_sw = (r_off - header_row - 1) < n_schweissteile
        for c, name in enumerate(FLAT_COLUMNS, 1):
            cell = ws.cell(r_off, c, entry.get(name))
            cell.border = BORDER
            if is_sw:
                cell.fill = SW_FILL

    auto_col_width(ws)
    return ws


# ── Blatt schreiben: Schlosserliste ──────────────────────────────────────────

def write_schlosser_sheet(wb, src_ws, header_row: int, aggregated: list, n_schweissteile: int, sheet_name: str):
    """Schlosserliste im Layout der Vorlage (38-02024-00_Schlosserliste)."""
    ws = wb.create_sheet(sheet_name)
    n_col = len(SCHLOSSER_HEADERS)
    copy_titelblock(src_ws, ws, header_row, n_col)

    # Titelblock für die Schlosserliste aus dem Quellblatt aufbauen.
    # Die vier Angaben aus I/J des Quellblatts werden an die vorgesehenen
    # Stellen des Titelblocks verschoben (siehe SCHLOSSER_MERGES: E4:F5 und
    # G4:G5 gehören zusammen - G3 lag außerhalb dieser Merge-Zeile, war ein Fehler).
    ws['F2'] = src_ws['I1'].value
    ws['H2'] = src_ws['J1'].value
    ws['I2'] = 'Rote Nr:'
    ws['F4'] = src_ws['I3'].value
    ws['H4'] = src_ws['J3'].value
    # Ursprüngliche I/J-Zellen aus dem kopierten Titelblock leeren - der Inhalt
    # ist jetzt nach F2/H2/F4/H4 verschoben, sonst stünde er doppelt da
    for coord in ('I1', 'J1', 'I3', 'J3'):
        ws[coord] = None

    # Merges ZUERST setzen - merge_cells() setzt nicht verankerte Zellen (z.B.
    # E5, F4, F5 innerhalb E4:F5) auf Default-Formatierung zurück, daher muss
    # die Schriftformatierung danach erfolgen, sonst geht sie dort verloren.
    for rng in SCHLOSSER_MERGES:
        try:
            ws.merge_cells(rng)
        except Exception:
            pass

    # Titelblock-Schrift: Bahnschrift SemiCondensed.
    for row in range(1, 7):
        for col in range(1, 12):
            cell = ws.cell(row, col)
            if row in (2, 3):
                cell.font = Font(name='Bahnschrift SemiCondensed', sz=16, bold=False)
            else:
                cell.font = Font(name='Bahnschrift SemiCondensed', sz=16, bold=True)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Wertbereich G4:G5 (Änderungsstand-Wert) groß und fett - NICHT E/F
    # (das ist das Label 'Änderungsstand:', siehe Ausnahmen unten)
    for row in range(4, 6):
        for col in (8, 9, 10, 11):  # H, I, J, K
            ws.cell(row, col).font = Font(name='Bahnschrift SemiCondensed', sz=24, bold=True)

    # Explizite Ausnahmen von der Standard-Logik oben:
    for coord in ('D2', 'E2'):
        ws[coord].font = Font(name='Bahnschrift SemiCondensed', sz=24, bold=True)
    for coord in ('D4', 'E4', 'D5', 'E5', 'F4', 'G4', 'F5', 'G5'):
        ws[coord].font = Font(name='Bahnschrift SemiCondensed', sz=16, bold=False)
    for coord in ('F2', 'F4'):
        ws[coord].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # Kopfzeile: Rot, Bahnschrift SemiCondensed 16 bold, weiße Schrift
    for c, (name, breite) in enumerate(zip(SCHLOSSER_HEADERS, SCHLOSSER_BREITEN), 1):
        cell = ws.cell(header_row, c, name)
        cell.font = Font(name='Bahnschrift SemiCondensed', sz=16, bold=True, color='FFFFFF')
        cell.fill = HDR_FILL_ROT
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = breite

    # Masse-Kopfzelle als Formel wie Vorlage (Stück jetzt Spalte C, Gewicht Spalte G)
    ws.cell(header_row, 7).value = '="Masse\n"&ROUNDUP(SUMPRODUCT($C$8:$C$999,$G$8:$G$999),0)&"kg"'

    ws.row_dimensions[header_row].height = 42

    # Datenzeilen: 16pt, Zeilenhöhe 42
    for r_off, entry in enumerate(aggregated, header_row + 1):
        hat_f = 'F' in str(entry.get('F,K,N') or '')

        dateiname = str(entry.get('Dateiname') or '').strip()
        benennung = str(entry.get('Benennung') or '').strip()
        ben_dat = f"{benennung}\n({dateiname})" if dateiname else benennung

        sachnr = str(entry.get('Sachnummer') or '').strip()
        norm = str(entry.get('Norm-Kurzbezeichnung') or '').strip()
        sac_norm = f"{sachnr}\n{norm}" if norm else sachnr

        mat = str(entry.get('Material') or '').strip()
        waerme = str(entry.get('Wärmebehandlung') or '').strip()
        mat_wae = f"{mat}\n{waerme}" if waerme and waerme not in ('-', '') else mat

        lief = str(entry.get('Lieferant') or '').strip()
        bem = str(entry.get('Bemerkungen') or '').strip()
        lief_bem = f"{lief}\n{bem}" if bem and bem not in ('-', '') else lief

        row_data = [
            None,  # Häkchen-Spalte: bleibt leer, zum manuellen Abhaken
            entry.get('Pos-Nr'),
            entry.get('Stück'),
            ben_dat,
            sac_norm,
            mat_wae,
            entry.get('Gewicht'),
            lief_bem,
            entry.get('F,K,N'),
            entry.get('V'),
            entry.get('Rev.'),
        ]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r_off, c, val)
            cell.font = SCHLOSSER_FONT
            # Pos rechtsbündig wie Vorlage, Häkchen-Spalte zentriert
            if c == 1:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            elif c == 2:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='right')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = BORDER
            if c == 7:
                cell.number_format = '0.00'
            # Alle Zeilen mit 'F' in Spalte H ('F,K,N') grau einfärben - auch
            # Schweißteile, die haben ja ebenfalls ein 'F' dort und brauchen
            # keine separate Sonderfarbe mehr.
            if hat_f:
                cell.fill = GRAU_FILL

        ws.row_dimensions[r_off].height = 42

    # A4-Drucklayout
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.2, footer=0.2)
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.oddFooter.left.text = "&A"
    ws.oddFooter.center.text = "Datum: &D"
    ws.oddFooter.right.text = "Seite &P von &N"
    ws.evenFooter.left.text = "&A"
    ws.evenFooter.center.text = "Datum: &D"
    ws.evenFooter.right.text = "Seite &P von &N"
    ws.oddFooter.left.size = 9
    ws.oddFooter.center.size = 9
    ws.oddFooter.right.size = 9
    ws.evenFooter.left.size = 9
    ws.evenFooter.center.size = 9
    ws.evenFooter.right.size = 9
    last_row = header_row + len(aggregated)
    ws.print_area = f"A1:{get_column_letter(n_col)}{last_row}"
    return ws


def _pdf_text(value) -> str:
    """Bereitet Text für eine PDF mit den eingebauten PDF-Schriften auf."""
    text = '' if value is None else str(value)
    text = text.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')
    return text.encode('cp1252', errors='replace').decode('latin-1')


def _export_schlosserliste_pdf_builtin(xlsx_path: Path, pdf_path: Path, sheet_name: str) -> None:
    """Erzeugt eine einfache druckbare PDF ohne Drittanbieter-Module."""
    _import_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Blatt '{sheet_name}' wurde für den PDF-Export nicht gefunden.")
    ws = wb[sheet_name]

    page_w, page_h = 595, 842
    left, right = 24, 571
    top, bottom = 810, 28
    col_widths = [34, 27, 112, 128, 84, 43, 70, 25, 25, 29]
    scale = (right - left) / sum(col_widths)
    col_widths = [w * scale for w in col_widths]
    x_positions = [left]
    for width in col_widths:
        x_positions.append(x_positions[-1] + width)

    rows = []
    for r in range(1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, 11)]
        if r <= 6 or r == 7 or any(v is not None for v in values):
            rows.append((r, values))

    masse = sum(float(ws.cell(r, 6).value) for r in range(8, ws.max_row + 1)
                if isinstance(ws.cell(r, 6).value, (int, float)))
    header = list(rows[[r for r, _ in rows].index(7)][1])
    header[5] = f"Masse\n{int(masse + 0.999999)}kg"
    rows = [(r, header if r == 7 else values) for r, values in rows]

    def lines(value, max_chars):
        raw = '' if value is None else str(value)
        result = []
        for part in raw.split('\n'):
            part = part or ' '
            while len(part) > max_chars:
                result.append(part[:max_chars])
                part = part[max_chars:]
            result.append(part)
        return result or [' ']

    pages = []
    page_rows = []
    y = top
    for row_no, values in rows:
        max_lines = max(len(lines(v, max(3, int(w / 4.5)))) for v, w in zip(values, col_widths))
        row_h = 22 if row_no <= 7 else max(22, min(48, 9 * max_lines + 7))
        if y - row_h < bottom and page_rows:
            pages.append(page_rows)
            page_rows = []
            y = top
        page_rows.append((row_no, values, y, row_h))
        y -= row_h
    if page_rows:
        pages.append(page_rows)

    objects = []
    page_refs = []
    today_text = date.today().strftime('%d.%m.%Y')
    for page_index, page in enumerate(pages, 1):
        commands = ['q', '0.75 w']
        for row_no, values, y, row_h in page:
            is_header = row_no == 7
            if is_header:
                commands += ['0.75 0 0 rg', f'{left:.2f} {y-row_h:.2f} {right-left:.2f} {row_h:.2f} re f', '1 1 1 rg']
            else:
                commands.append('0 0 0 rg')
            commands.append('0 0 0 RG')
            for x in x_positions:
                commands.append(f'{x:.2f} {y-row_h:.2f} m {x:.2f} {y:.2f} l')
            commands.append(f'{left:.2f} {y-row_h:.2f} m {right:.2f} {y-row_h:.2f} l')
            commands.append(f'{left:.2f} {y:.2f} m {right:.2f} {y:.2f} l S')
            for c, (value, x0, x1) in enumerate(zip(values, x_positions[:-1], x_positions[1:])):
                max_chars = max(3, int((x1 - x0 - 4) / (5.0 if is_header else 4.5)))
                for line_no, text in enumerate(lines(value, max_chars)[:5]):
                    tx = x0 + 2
                    ty = y - 12 - line_no * 8
                    commands += [f'BT /F1 {6.5 if is_header else 6} Tf {tx:.2f} {ty:.2f} Td ({_pdf_text(text)}) Tj ET']
            commands.append('S')
        commands += [
            '0 0 0 rg',
            f'BT /F1 7 Tf {left:.2f} 14 Td ({_pdf_text(sheet_name)}) Tj ET',
            f'BT /F1 7 Tf {page_w/2-25:.2f} 14 Td (Datum: {_pdf_text(today_text)}) Tj ET',
            f'BT /F1 7 Tf {right-95:.2f} 14 Td (Seite {page_index} von {len(pages)}) Tj ET',
        ]
        commands.append('Q')
        stream = '\n'.join(commands).encode('latin-1', errors='replace')
        objects.append(stream)

    pdf = bytearray(b'%PDF-1.4\n%\xe2\xe3\xcf\xd3\n')
    offsets = [0]
    def add(obj):
        offsets.append(len(pdf))
        pdf.extend(obj)

    add(b'1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n')
    kids = ' '.join(f'{4 + i * 2} 0 R' for i in range(len(objects)))
    add(f'2 0 obj << /Type /Pages /Kids [{kids}] /Count {len(objects)} >> endobj\n'.encode('ascii'))
    add(b'3 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >> endobj\n')
    for i, stream in enumerate(objects):
        page_id = 4 + i * 2
        content_id = page_id + 1
        add(f'{page_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] /Resources << /Font << /F1 3 0 R >> >> /Contents {content_id} 0 R >> endobj\n'.encode('ascii'))
        add(f'{content_id} 0 obj << /Length {len(stream)} >> stream\n'.encode('ascii') + stream + b'\nendstream endobj\n')
    xref = len(pdf)
    pdf.extend(f'xref\n0 {len(offsets)}\n0000000000 65535 f \n'.encode('ascii'))
    for off in offsets[1:]:
        pdf.extend(f'{off:010d} 00000 n \n'.encode('ascii'))
    pdf.extend(f'trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n'.encode('ascii'))
    pdf_path.write_bytes(pdf)
    wb.close()


def _export_via_excel(xlsx_path: Path, pdf_path: Path, sheet_name: str, header_row: int) -> bool:
    """Exportiert das Blatt 1:1 wie in Excel als PDF, per Excel-COM-Automatisierung
    (über ein kurzes PowerShell-Skript). Das ist die einzige Methode, die
    garantiert exakt wie die echte Excel-Formatierung aussieht (Merges, Fonts,
    Spaltenbreiten, Seitenumbrüche) - ein selbstgebauter PDF-Renderer kann das
    nur annähern. Erfordert Windows + installiertes Excel; gibt sonst False
    zurück, damit der Aufrufer auf den eingebauten Renderer zurückfallen kann.

    PrintTitleRows (Kopfzeile auf jeder Seite wiederholen) wird hier NOCHMAL
    explizit über COM gesetzt, nicht nur aus der Datei übernommen - Excel
    respektiert die per openpyxl geschriebene Print_Titles-Definition beim
    automatisierten ExportAsFixedFormat nicht immer zuverlässig."""
    if sys.platform != 'win32':
        return False

    ps_code = r'''
param([string]$XlsxPfad, [string]$PdfPfad, [string]$Blatt, [int]$KopfZeile)
$Excel = $null; $Wb = $null; $Ws = $null
try {
    $Excel = New-Object -ComObject Excel.Application
    $Excel.Visible = $false
    $Excel.DisplayAlerts = $false
    $Wb = $Excel.Workbooks.Open($XlsxPfad)
    $Ws = $Wb.Sheets.Item($Blatt)
    $Ws.Select()
    $Ws.PageSetup.PrintTitleRows = "`$$KopfZeile`:`$$KopfZeile"
    $Wb.ActiveSheet.ExportAsFixedFormat(0, $PdfPfad)  # 0 = xlTypePDF
} catch {
    Write-Error $_.Exception.Message
    exit 1
} finally {
    if ($Wb) { $Wb.Close($false) }
    if ($Excel) { $Excel.Quit() }
    foreach ($o in @($Ws, $Wb, $Excel)) {
        if ($o) { [System.Runtime.Interopservices.Marshal]::ReleaseComObject($o) | Out-Null }
    }
}
'''
    import tempfile
    with tempfile.NamedTemporaryFile(mode='w', suffix='.ps1', delete=False, encoding='utf-8') as fh:
        fh.write(ps_code)
        ps1_path = fh.name

    try:
        result = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-File', ps1_path,
             str(xlsx_path), str(pdf_path), sheet_name, str(header_row)],
            capture_output=True, text=True, timeout=120,
        )
    except FileNotFoundError:
        return False  # kein PowerShell verfügbar
    finally:
        try:
            Path(ps1_path).unlink()
        except OSError:
            pass

    if result.returncode != 0:
        log('WARN', f"Excel-PDF-Export fehlgeschlagen: {(result.stderr or result.stdout).strip()}")
        return False
    return pdf_path.exists()


def export_schlosserliste_pdf(xlsx_path: Path, pdf_path: Path, sheet_name: str, header_row: int) -> None:
    """Exportiert die Schlosserliste als PDF - bevorzugt per Excel (exakte
    Formatierung, Kopfzeile wiederholt sich zuverlässig auf jeder Seite),
    sonst per eingebautem Not-Renderer (Näherung)."""
    if _export_via_excel(xlsx_path, pdf_path, sheet_name, header_row):
        log('INFO', f"PDF exportiert (Excel): {pdf_path.name}")
        return
    log('WARN', "Excel-Export nicht möglich - falle zurück auf eingebauten PDF-Renderer (Näherung, keine 1:1-Formatierung)")
    _export_schlosserliste_pdf_builtin(xlsx_path, pdf_path, sheet_name)
    log('INFO', f"PDF exportiert (eingebauter Renderer): {pdf_path.name}")


def _reorder_sheets(wb, gfu_name: str, schlosser_name: str, latest_stand_name: str) -> None:
    """Stellt sicher: [GFU, Schlosserliste, latest Stand, rest...]"""
    order = []
    if gfu_name in wb.sheetnames:
        order.append(gfu_name)
    if schlosser_name in wb.sheetnames:
        order.append(schlosser_name)
    if latest_stand_name in wb.sheetnames:
        order.append(latest_stand_name)
    for n in wb.sheetnames:
        if n not in order:
            order.append(n)
    wb._sheets = [wb[n] for n in order]


# ── Ordner-Erkennung ──────────────────────────────────────────────────────────

def find_csv(ordner: Path) -> Path:
    """Sucht im Ordner nach genau einer CSV-Datei."""
    csvs = [f for f in ordner.iterdir() if f.suffix.lower() == '.csv' and not f.name.startswith('~$')]
    if len(csvs) == 0:
        raise FileNotFoundError("Keine CSV-Datei im Ordner gefunden.")
    if len(csvs) > 1:
        raise ValueError("Mehrere CSV-Dateien gefunden – bitte nur eine ablegen:\n" + "\n".join(f"  {f.name}" for f in csvs))
    return csvs[0]


# ── Hauptverarbeitung ─────────────────────────────────────────────────────────

def process(ordner: Path, interaktiv: bool = True) -> int:
    print("=" * 62)
    print(f"[W] GFU Stücklisten-Generator und Validator {SCRIPT_VERSION}")
    print("=" * 62)

    try:
        csv_path = find_csv(ordner)
    except (FileNotFoundError, ValueError) as e:
        log('FEHLER', str(e))
        return 1

    print(f"Ordner : {ordner}")
    print(f"Quelle : {csv_path.name}")

    try:
        _, src_ws, _ = load_csv(csv_path)
    except Exception as e:
        log('FEHLER', f"CSV konnte nicht gelesen werden: {e}")
        return 1

    try:
        header_row, hmap = find_header(src_ws)
    except ValueError as e:
        log('FEHLER', str(e))
        return 1

    auftragsnr = find_auftragsnummer(src_ws, header_row)
    xlsx_path = ordner / f"{auftragsnr}_STUECKLISTE.XLSX"
    modus = 'FORTSETZEN' if xlsx_path.exists() else 'NEU'
    print(f"Modus  : {modus}" + (f"  ({xlsx_path.name} gefunden)" if modus == 'FORTSETZEN' else ""))
    print("-" * 62)
    print(f"Archivnummer: {auftragsnr}")
    print("-" * 62)

    rows = read_rows(src_ws, header_row, hmap)
    global ZEILEN_BREITE
    if rows:
        ZEILEN_BREITE = max(3, len(str(max(e['row'] for e in rows))))

    valid, children, parent_of = build_tree(rows)

    problems = validate_all(rows, valid, children)
    had_error = any(l == 'FEHLER' for _, l, _, _, _, _ in problems)
    probleme_je_zeile: dict = {}
    for z, l, m, b, s, p in problems:
        log(l, m, zeile=z, benennung=b, sachnummer=s, posnr=p)
        probleme_je_zeile.setdefault(z, []).append(f"[{_LEVEL_SYMBOL[l]}] {m}")

    schweissteil_rows = find_und_resolve_schweissteil(valid, children, interaktiv=interaktiv)

    schweissteile, sonstige = collect_items(valid, children, parent_of, schweissteil_rows)
    aggregated, n_sw = aggregate_items(schweissteile, sonstige)
    print()
    log('INFO', f"GFU-Liste: {len(aggregated)} Positionen ({n_sw} Schweißteile vorne, {len(aggregated) - n_sw} weitere)")

    heute = date.today().strftime('%Y%m%d')
    stand_name = f"{auftragsnr}_Stand_{heute}"[:31]
    gfu_name = f"{auftragsnr}_fuer_GFU"[:31]
    schlosser_name = f"{auftragsnr}_SCHLOSSERLISTE"[:31]

    if xlsx_path.exists():
        _import_openpyxl()
        wb = openpyxl.load_workbook(xlsx_path)
        for name in [gfu_name, schlosser_name]:
            if name in wb.sheetnames:
                del wb[name]
                log('INFO', f"Blatt '{name}' gelöscht (wird neu erzeugt)")
    else:
        _import_openpyxl()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    write_source_sheet(wb, src_ws, header_row, hmap, valid, stand_name, probleme_je_zeile)
    log('INFO', f"Aktuelle Liste: '{stand_name}'")

    write_flat_sheet(wb, src_ws, header_row, aggregated, n_sw, gfu_name)
    log('INFO', f"GFU-Blatt: '{gfu_name}'")

    write_schlosser_sheet(wb, src_ws, header_row, aggregated, n_sw, schlosser_name)
    log('INFO', f"Schlosserliste: '{schlosser_name}'")

    _reorder_sheets(wb, gfu_name, schlosser_name, stand_name)

    try:
        wb.save(xlsx_path)
        print(f"\n[>] Gespeichert: {xlsx_path.name}")
    except Exception as e:
        log('FEHLER', f"XLSX konnte nicht gespeichert werden: {e}")
        return 1

    # PDF heißt wie das Schlosserliste-Tabellenblatt
    pdf_path = ordner / f"{schlosser_name}.PDF"
    try:
        export_schlosserliste_pdf(xlsx_path, pdf_path, schlosser_name, header_row)
    except Exception as e:
        log('FEHLER', f"PDF konnte nicht exportiert werden: {e}")
        return 1

    try:
        csv_path.unlink()
        log('INFO', f"CSV gelöscht: {csv_path.name}")
    except Exception as e:
        log('WARN', f"CSV konnte nicht gelöscht werden: {e}")

    print("=" * 62)
    return 1 if had_error else 0


def main() -> None:
    parser = argparse.ArgumentParser(description="GFU Stücklisten-Generator – konvertiert Creo-CSV in strukturierte XLSX.")
    parser.add_argument("ordner", help="Ordner mit genau einer CSV-Quelldatei")
    parser.add_argument("--batch", action="store_true", help="Kein interaktiver Modus: Schweißteil-Kandidaten werden automatisch bestätigt")
    args = parser.parse_args()

    ordner = Path(args.ordner)
    if not ordner.is_dir():
        print(f"[X]   Ordner nicht gefunden: {ordner}")
        sys.exit(1)

    rc = process(ordner, interaktiv=not args.batch)
    sys.exit(rc)


if __name__ == "__main__":
    main()